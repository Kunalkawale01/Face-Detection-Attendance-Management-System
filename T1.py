from flask import Flask, request
import os, cv2, face_recognition
from openpyxl import Workbook, load_workbook

import datetime

# Ask for the attendance date at the start
attendance_date = input("Enter attendance date (YYYY-MM-DD): ")
try:
    datetime.datetime.strptime(attendance_date, "%Y-%m-%d")
except ValueError:
    print("Invalid date format! Use YYYY-MM-DD.")
    exit(1)

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
KNOWN_FACES_FOLDER = 'known_faces'
EXCEL_FILE = 'attendance.xlsx'

# Ensure folders exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Load known face encodings
known_encodings = []
known_names = []

for file in os.listdir(KNOWN_FACES_FOLDER):
    img = face_recognition.load_image_file(f"{KNOWN_FACES_FOLDER}/{file}")
    enc = face_recognition.face_encodings(img)
    if enc:
        known_encodings.append(enc[0])
        known_names.append(file.split(".")[0])

# Create Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name"])  # Only Name initially, dates will be added
    wb.save(EXCEL_FILE)

@app.route("/upload", methods=["POST"])
def upload_image():
    file = request.data
    file_path = os.path.join(UPLOAD_FOLDER, "incoming.jpg")
    with open(file_path, "wb") as f:
        f.write(file)

    img = face_recognition.load_image_file(file_path)
    face_locations = face_recognition.face_locations(img)
    face_encodings = face_recognition.face_encodings(img, face_locations)

    for face_encoding in face_encodings:
        matches = face_recognition.compare_faces(known_encodings, face_encoding)
        if True in matches:
            match_index = matches.index(True)
            name = known_names[match_index]
            log_attendance(name, attendance_date)

    return "OK", 200

def log_attendance(name, date):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Ensure date column exists
    if date not in [cell.value for cell in ws[1]]:
        ws.cell(row=1, column=ws.max_column + 1).value = date

    date_col = [cell.value for cell in ws[1]].index(date) + 1

    # Search for the student row
    student_row = None
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == name:
            student_row = i
            break

    # Add student if not found
    if not student_row:
        student_row = ws.max_row + 1
        ws.cell(row=student_row, column=1).value = name

    # Only mark if not already marked
    if not ws.cell(row=student_row, column=date_col).value:
        ws.cell(row=student_row, column=date_col).value = "P"
        print(f"{name} marked present on {date}")
    else:
        print(f"{name} already marked present on {date}.")

    wb.save(EXCEL_FILE)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
